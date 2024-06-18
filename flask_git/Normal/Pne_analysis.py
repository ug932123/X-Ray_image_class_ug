import os
import tensorflow as tf
from keras.utils import load_img, img_to_array
import numpy as np
import openpyxl
from sklearn.metrics import confusion_matrix, accuracy_score, precision_score, recall_score, f1_score
import matplotlib.pyplot as plt
import seaborn as sns

# Load the model
model_efficientnetb7 = tf.keras.models.load_model("data/Final_Efficientb7_model.h5")
model_efficientnetb7.load_weights("data/Final_Efficientb7_weights.h5")

categories = ['Normal', 'Pneumonia']
true_labels = []
predicted_labels = []
predictions = []

file_path_sheet = 'x.xlsx'
sheet_name = 'Pneumonia'

def append_to_excel(file_path_sheet, sheet_name, data):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_path_sheet)
    
    # Check if the sheet exists, if not, create it
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    
    # Append the data to the sheet
    for row in data:
        sheet.append(row)
    
    # Save the workbook
    workbook.save(file_path_sheet)

for i in range(1, 21):
    #file_path = f'Normal/l{i}.jpg'
    file_path = f'Pneumonia/p{i}.jpg'
    
    try:
        test_image = load_img(file_path, target_size=(224, 224))
        img_array = img_to_array(test_image)
        img_array = np.expand_dims(img_array, axis=0)
        
        prediction = model_efficientnetb7.predict(img_array)
        predictions.append(prediction)
        predicted_label = np.argmax(prediction)
        predicted_labels.append(predicted_label)
        
        confidence = np.array(prediction)
        if confidence[0][0] > confidence[0][1]:
            output_confidence = confidence[0][0]
        else:
            output_confidence = confidence[0][1]
        output_confidence = output_confidence * 100
        output_confidence = round(output_confidence, 2)
        
        true_label = 1  # Assuming that all images in 'Pneumonia' folder are labeled as 'Pneumonia'
        true_labels.append(true_label)
        
        print(categories[predicted_label])
        print(output_confidence)
        print(f'Pneumonia/p{i}.jpg: {prediction}')
        
        data_to_append = [[f'p{i}.jpg', categories[predicted_label], output_confidence]]
        append_to_excel(file_path_sheet, sheet_name, data_to_append)
    
    except Exception as e:
        print(f'Error processing file {file_path}: {e}')

# Calculate metrics
conf_matrix = confusion_matrix(true_labels, predicted_labels)
accuracy = accuracy_score(true_labels, predicted_labels)
precision = precision_score(true_labels, predicted_labels)
recall = recall_score(true_labels, predicted_labels)
f1 = f1_score(true_labels, predicted_labels)

print(f'Confusion Matrix:\n{conf_matrix}')
print(f'Accuracy: {accuracy:.2f}')
print(f'Precision: {precision:.2f}')
print(f'Recall: {recall:.2f}')
print(f'F1 Score: {f1:.2f}')

# Save metrics to Excel
metrics_data = [
    ['Metric', 'Value'],
    ['Accuracy', accuracy],
    ['Precision', precision],
    ['Recall', recall],
    ['F1 Score', f1]
]
append_to_excel(file_path_sheet, 'Metrics_Pne', metrics_data)

# Plot confusion matrix
plt.figure(figsize=(10, 8))
sns.heatmap(conf_matrix, annot=True, fmt='d', cmap='Blues', xticklabels=categories, yticklabels=categories)
plt.ylabel('True Label')
plt.xlabel('Predicted Label')
plt.title('Confusion Matrix For Pneumonia Test Data')
plt.savefig('Pne_confusion_matrix.png')
plt.show()

